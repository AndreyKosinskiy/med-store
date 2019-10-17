import openpyxl
from django.http import HttpResponse
import time
DOC_TYPE_LIST = ('+','-')
DOC_TYPE_CELL = 'B1'
SIZE_OF_HEAD = 3

def get_info_from_excel(file):
    work_book = openpyxl.load_workbook(file)
    sheet_list = work_book.sheetnames
    work_sheet = work_book.active #get active sheet 
    
    doc_type = work_sheet[DOC_TYPE_CELL].value
    rows = work_sheet.rows
    good_table = tuple(rows)[SIZE_OF_HEAD:]
    return  {
                'doc_type':doc_type,
                'good_table':good_table
            }


def is_valid_or_list_error(doc):
    def analyzer_doc_type(value):
        return  value in DOC_TYPE_LIST
    pull_erorrs = {
        f'doc_type_error (cell {DOC_TYPE_CELL}):':analyzer_doc_type(doc['doc_type']),
        'empty_file:':(doc is not None)
        }
    if False in pull_erorrs.values():
        return pull_erorrs
    else:
        return True

def print_table(table):
    for row in table:
        for cell in row:
            print (cell,end='')
        print ()

def build_book(data_range,table):
    wb = openpyxl.Workbook()
    ws = wb.active
    #ws.merge_cells('A1:A3')
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['C'].width = 15
    start = data_range[0].date().strftime("%d/%m/%Y")
    end = data_range[1].date().strftime("%d/%m/%Y")
    ws['A1'] = f"Срез c {start} по {end}"
    ws['A2'] = "Название"
    ws['B2'] = "Партия"
    ws['C2'] = "Количество"
    min_row=3
    max_col=1000
    max_row=len(table)+min_row
    for row,row_table in zip(ws.iter_rows(min_row=min_row, max_col=max_col, max_row=max_row),table):
        for cell,cell_table in  zip(row,row_table):
            cell.value = cell_table
            print(cell.value)
    stream = openpyxl.writer.excel.save_virtual_workbook(wb)
    response = HttpResponse(stream, content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="foo.xlsx"'
    return response
if __name__ == "__main__":   
    document = get_info_from_excel('C:\\Users\\Admin\\Desktop\\put_file.xlsx')
    analyzer_doc(document)


    
